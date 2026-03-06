"""
Excel 数据处理核心类
用于读取、分析、处理 Excel 数据，生成报告和项目表单
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


class ExcelProcessor:
    """Excel 数据处理工具类"""

    def __init__(self):
        pass

    def read_excel(self, filepath, sheet_name=0, **kwargs):
        """
        读取 Excel 文件

        参数:
            filepath: 文件路径
            sheet_name: 工作表名称或索引
            **kwargs: pandas read_excel 的其他参数

        返回:
            DataFrame
        """
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, **kwargs)
            print(f"✓ 成功读取 {filepath}")
            print(f"  - 工作表: {sheet_name}")
            print(f"  - 数据行数: {len(df)}")
            print(f"  - 列数: {len(df.columns)}")
            return df
        except Exception as e:
            print(f"✗ 读取失败: {e}")
            return None

    def get_sheet_names(self, filepath):
        """获取 Excel 文件中所有工作表名称"""
        try:
            xlsx = pd.ExcelFile(filepath)
            return xlsx.sheet_names
        except Exception as e:
            print(f"✗ 获取工作表名称失败: {e}")
            return []

    def write_excel(self, df, filepath, sheet_name='Sheet1', index=False, **kwargs):
        """
        写入 Excel 文件

        参数:
            df: DataFrame
            filepath: 输出文件路径
            sheet_name: 工作表名称
            index: 是否写入索引
        """
        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            df.to_excel(filepath, sheet_name=sheet_name, index=index, **kwargs)
            print(f"✓ 成功写入 {filepath}")
            return True
        except Exception as e:
            print(f"✗ 写入失败: {e}")
            return False

    def format_excel(self, filepath, sheet_name='Sheet1', header_color='366092'):
        """
        格式化 Excel 表格（标题加粗、自动列宽、添加边框）

        参数:
            filepath: 文件路径
            sheet_name: 工作表名称
            header_color: 标题背景色
        """
        try:
            book = load_workbook(filepath)
            sheet = book[sheet_name]

            # 设置标题样式
            for cell in sheet[1]:
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 自动列宽
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            # 中文宽度调整
                            chinese_chars = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                            cell_length = cell_length + chinese_chars
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = max(max_length + 2, 8) * 1.1
                sheet.column_dimensions[column].width = adjusted_width

            # 添加边框
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            for row in sheet.iter_rows(min_row=1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 冻结首行
            sheet.freeze_panes = 'A2'

            book.save(filepath)
            print(f"✓ 格式化完成: {filepath}")
            return True
        except Exception as e:
            print(f"✗ 格式化失败: {e}")
            return False

    def merge_excels(self, file_list, output_file, add_source=False):
        """
        合并多个 Excel 文件

        参数:
            file_list: 文件列表
            output_file: 输出文件路径
            add_source: 是否添加来源文件列
        """
        try:
            all_dfs = []
            for file in file_list:
                df = pd.read_excel(file)
                if add_source:
                    df['来源文件'] = os.path.basename(file)
                all_dfs.append(df)

            merged = pd.concat(all_dfs, ignore_index=True)
            merged.to_excel(output_file, index=False)
            print(f"✓ 合并完成: {len(file_list)} 个文件 → {output_file}")
            return merged
        except Exception as e:
            print(f"✗ 合并失败: {e}")
            return None

    def create_pivot_table(self, df, index, columns=None, values=None, aggfunc='sum', fill_value=0):
        """
        创建数据透视表

        参数:
            df: DataFrame
            index: 行字段
            columns: 列字段
            values: 值字段
            aggfunc: 聚合函数 (sum, mean, count, max, min)
            fill_value: 填充值

        返回:
           透视表 DataFrame
        """
        try:
            pivot = df.pivot_table(
                index=index,
                columns=columns,
                values=values,
                aggfunc=aggfunc,
                fill_value=fill_value
            )
            return pivot
        except Exception as e:
            print(f"✗ 创建透视表失败: {e}")
            return None

    def export_multiple_sheets(self, data_dict, output_file, format_all=False):
        """
        导出多个 sheet 到一个 Excel 文件

        参数:
            data_dict: 字典 {sheet_name: df}
            output_file: 输出文件路径
            format_all: 是否格式化所有 sheet
        """
        try:
            with pd.ExcelWriter(output_file) as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            if format_all:
                for sheet_name in data_dict.keys():
                    self.format_excel(output_file, sheet_name=sheet_name)

            print(f"✓ 导出完成: {len(data_dict)} 个工作表 → {output_file}")
            return True
        except Exception as e:
            print(f"✗ 导出失败: {e}")
            return False

    def analyze_data(self, df, numeric_only=True):
        """
        数据分析摘要

        参数:
            df: DataFrame
            numeric_only: 是否只分析数值列

        返回:
            分析结果字典
        """
        analysis = {
            'shape': df.shape,
            'columns': list(df.columns),
            'dtypes': df.dtypes.to_dict(),
            'null_counts': df.isnull().sum().to_dict(),
            'memory_usage': f"{df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB"
        }

        if numeric_only:
            numeric_df = df.select_dtypes(include=[pd.np.number])
            if not numeric_df.empty:
                analysis['describe'] = numeric_df.describe().to_dict()

        return analysis

    def clean_data(self, df, dropna_cols=None, fillna_dict=None, strip_cols=None):
        """
        数据清洗

        参数:
            df: DataFrame
            dropna_cols: 删除空值的列列表
            fillna_dict: 填充字典 {列名: 填充值}
            strip_cols: 去除空格的列列表

        返回:
            清洗后的 DataFrame
        """
        df_clean = df.copy()

        # 删除空值
        if dropna_cols:
            df_clean = df_clean.dropna(subset=dropna_cols)

        # 填充空值
        if fillna_dict:
            df_clean = df_clean.fillna(fillna_dict)

        # 去除字符串空格
        if strip_cols:
            for col in strip_cols:
                if col in df_clean.columns:
                    df_clean[col] = df_clean[col].astype(str).str.strip()

        return df_clean


class ProjectManager:
    """项目管理工具类"""

    def __init__(self):
        self.processor = ExcelProcessor()

    def create_progress_tracker(self, output_file='project_progress.xlsx'):
        """创建项目进度跟踪表"""
        data = {
            '任务ID': [f'T{i:03d}' for i in range(1, 6)],
            '任务名称': ['需求分析', '系统设计', '开发实施', '测试验证', '部署上线'],
            '负责人': ['张三', '李四', '王五', '赵六', '张三'],
            '开始日期': [datetime.now().date() - timedelta(days=5)] * 5,
            '计划完成': [datetime.now().date() + timedelta(days=i*5+5) for i in range(5)],
            '实际完成': [None] * 5,
            '状态': ['进行中', '待开始', '待开始', '待开始', '待开始'],
            '进度(%)': [30, 0, 0, 0, 0],
            '优先级': ['高', '高', '中', '中', '高'],
            '备注': [''] * 5
        }

        df = pd.DataFrame(data)
        self.processor.write_excel(df, output_file)
        self.processor.format_excel(output_file)
        return df

    def create_resource_allocation(self, output_file='resource_allocation.xlsx'):
        """创建资源分配表"""
        data = {
            '资源ID': [f'R{i:03d}' for i in range(1, 6)],
            '资源名称': ['开发工程师A', '开发工程师B', '测试工程师', 'UI设计师', '项目经理'],
            '技能': ['Python', 'Java', '自动化测试', 'Figma', '敏捷管理'],
            '可用性(%)': [80, 100, 60, 70, 90],
            '当前项目': ['项目A', '项目A', '项目B', '项目A', '项目A,项目B'],
            '成本/小时': [200, 180, 150, 220, 300],
            '联系方式': ['email1@company.com', 'email2@company.com', 'email3@company.com',
                        'email4@company.com', 'email5@company.com']
        }

        df = pd.DataFrame(data)
        self.processor.write_excel(df, output_file)
        self.processor.format_excel(output_file)
        return df

    def create_risk_register(self, output_file='risk_register.xlsx'):
        """创建风险登记册"""
        data = {
            '风险ID': [f'RISK{i:03d}' for i in range(1, 4)],
            '风险描述': ['技术选型不成熟', '关键人员离职', '需求变更频繁'],
            '风险类别': ['技术风险', '人员风险', '需求风险'],
            '可能性(1-5)': [3, 2, 4],
            '影响程度(1-5)': [4, 5, 3],
            '风险等级': ['高', '中', '高'],
            '应对策略': ['增加技术调研', '加强文档建设', '建立变更流程'],
            '责任人': ['技术总监', '项目经理', '产品经理'],
            '状态': ['监控中', '已缓解', '监控中'],
            '更新日期': [datetime.now().date()] * 3
        }

        df = pd.DataFrame(data)
        self.processor.write_excel(df, output_file)
        self.processor.format_excel(output_file)
        return df

    def generate_status_report(self, progress_file, output_file='project_report.md'):
        """
        生成项目状态报告

        参数:
            progress_file: 进度表文件路径
            output_file: 输出报告文件路径

        返回:
            报告文本
        """
        df = self.processor.read_excel(progress_file)
        if df is None:
            return None

        report = []
        report.append("# 📊 项目状态报告")
        report.append(f"\n生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # 项目概览
        total_tasks = len(df)
        completed = len(df[df['状态'] == '已完成'])
        in_progress = len(df[df['状态'] == '进行中'])
        not_started = len(df[df['状态'] == '待开始'])
        avg_progress = df['进度(%)'].mean() if '进度(%)' in df.columns else 0

        report.append("\n## 📈 项目概览")
        report.append(f"- **总任务数**: {total_tasks}")
        report.append(f"- **已完成**: {completed} ({completed/total_tasks*100:.1f}%)")
        report.append(f"- **进行中**: {in_progress} ({in_progress/total_tasks*100:.1f}%)")
        report.append(f"- **待开始**: {not_started} ({not_started/total_tasks*100:.1f}%)")
        report.append(f"- **整体进度**: {avg_progress:.1f}%")

        # 任务列表
        report.append("\n## 📋 任务详情")
        for _, row in df.iterrows():
            status_icon = {'已完成': '✅', '进行中': '🔄', '待开始': '⏳'}.get(row['状态'], '❓')
            report.append(f"\n{status_icon} **{row['任务名称']}**")
            report.append(f"   - ID: {row['任务ID']} | 状态: {row['状态']} | 进度: {row['进度(%)']}%")
            report.append(f"   - 负责人: {row['负责人']} | 优先级: {row['优先级']}")
            if '计划完成' in row and row['计划完成']:
                report.append(f"   - 计划完成: {row['计划完成'].strftime('%Y-%m-%d')}")
            if pd.notna(row.get('备注')) and row['备注']:
                report.append(f"   - 备注: {row['备注']}")

        # 延期任务
        if '计划完成' in df.columns:
            delayed = df[(df['状态'] != '已完成') &
                        (pd.to_datetime(df['计划完成']) < datetime.now())]
            if not delayed.empty:
                report.append("\n## ⚠️ 延期任务")
                for _, row in delayed.iterrows():
                    report.append(f"- **{row['任务名称']}** - 计划: {row['计划完成'].strftime('%Y-%m-%d')} - 负责人: {row['负责人']}")

        # 按负责人统计
        if '负责人' in df.columns:
            report.append("\n## 👥 负责人任务分布")
            for person in df['负责人'].unique():
                person_tasks = df[df['负责人'] == person]
                report.append(f"- **{person}**: {len(person_tasks)} 个任务 (平均进度: {person_tasks['进度(%)'].mean():.1f}%)")

        # 保存报告
        report_text = '\n'.join(report)
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(report_text)

        print(f"✓ 报告生成: {output_file}")
        return report_text

    def generate_weekly_report(self, progress_file, start_date, end_date):
        """
        生成周报

        参数:
            progress_file: 进度表文件路径
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)

        返回:
            周报文本
        """
        df = self.processor.read_excel(progress_file)
        if df is None:
            return None

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        report = []
        report.append("# 📅 项目周报")
        report.append(f"\n周期: {start} ~ {end}")
        report.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # 本周完成的任务
        if '实际完成' in df.columns:
            completed_dates = pd.to_datetime(df['实际完成'])
            completed_this_week = df[
                (df['状态'] == '已完成') &
                (completed_dates >= pd.Timestamp(start)) &
                (completed_dates <= pd.Timestamp(end))
            ]

            report.append(f"\n## ✅ 本周完成 ({len(completed_this_week)} 个任务)")
            for _, task in completed_this_week.iterrows():
                report.append(f"- **{task['任务名称']}** - 负责人: {task['负责人']} - 完成时间: {task['实际完成'].strftime('%Y-%m-%d')}")

        # 本周新增的任务
        if '开始日期' in df.columns:
            started_dates = pd.to_datetime(df['开始日期'])
            started_this_week = df[
                (started_dates >= pd.Timestamp(start)) &
                (started_dates <= pd.Timestamp(end))
            ]

            report.append(f"\n## 🆕 本周新增 ({len(started_this_week)} 个任务)")
            for _, task in started_this_week.iterrows():
                report.append(f"- **{task['任务名称']}** - 负责人: {task['负责人']} - 优先级: {task['优先级']}")

        # 进行中的任务
        in_progress = df[df['状态'] == '进行中']
        if not in_progress.empty:
            report.append(f"\n## 🔄 进行中 ({len(in_progress)} 个任务)")
            for _, task in in_progress.iterrows():
                report.append(f"- **{task['任务名称']}** - 进度: {task['进度(%)']}% - 负责人: {task['负责人']}")

        return '\n'.join(report)


if __name__ == '__main__':
    # 测试代码
    print("=== Excel 数据处理工具 ===\n")

    # 创建项目管理器
    pm = ProjectManager()

    # 创建模板文件
    print("1. 创建项目管理模板...")
    pm.create_progress_tracker('templates/进度表.xlsx')
    pm.create_resource_allocation('templates/资源表.xlsx')
    pm.create_risk_register('templates/风险表.xlsx')

    # 生成报告
    print("\n2. 生成项目状态报告...")
    report = pm.generate_status_report('templates/进度表.xlsx', 'templates/项目状态报告.md')
    print("\n" + report[:200] + "...")

    print("\n✓ 测试完成!")
