"""
项目风险管理工具
用于风险识别、评估、跟踪和复盘
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json


class RiskManager:
    """风险管理工具类"""

    def __init__(self):
        pass

    def create_risk_register(self, output_file='风险登记册.xlsx'):
        """
        创建风险登记册

        字段说明：
        - 风险ID: 唯一标识
        - 风险描述: 清晰描述风险
        - 风险类别: 技术/资源/时间/范围/质量/外部
        - 可能性(1-3): 1=低, 2=中, 3=高
        - 影响程度(1-3): 1=低, 2=中, 3=高
        - 风险等级: 可能性×影响程度 (1-3低, 4-6中, 7-9高)
        - 应对策略: 规避/缓解/转移/接受
        - 应对措施: 具体措施
        - 责任人: 负责跟踪的人
        - 状态: 监控中/已缓解/已关闭/已触发
        - 触发事件: 如已触发，描述事件
        - 更新日期: 最后更新时间
        """
        data = {
            '风险ID': [f'R{i:03d}' for i in range(1, 11)],
            '风险描述': [
                '关键技术选型不成熟，可能导致项目延期',
                '关键开发人员可能离职，影响项目进度',
                '需求变更频繁，导致范围蔓延',
                '第三方API服务不稳定，可能影响系统',
                '测试资源不足，质量可能受影响',
                '性能要求高，技术难点多',
                '多系统集成，兼容性问题',
                '外包团队配合度不高，可能延期',
                '数据迁移风险，可能导致数据丢失',
                '干系人期望不一致，验收困难'
            ],
            '风险类别': ['技术', '资源', '范围', '技术', '资源',
                        '技术', '技术', '资源', '技术', '外部'],
            '可能性(1-3)': [1, 2, 3, 2, 2, 2, 2, 3, 2, 2],
            '影响程度(1-3)': [3, 3, 2, 2, 2, 3, 3, 2, 3, 2],
            '风险等级': ['低', '中', '中', '中', '中', '中', '中', '中', '中', '中'],
            '应对策略': ['缓解', '缓解', '缓解', '缓解', '缓解',
                         '缓解', '缓解', '转移', '缓解', '缓解'],
            '应对措施': [
                '提前进行技术调研（2周），做POC验证',
                '建立知识文档，人员备份，技术交叉培训',
                '建立变更流程，变更影响评估，优先级排序',
                '做接口mock，监控，准备备选方案',
                '自动化测试，交叉测试',
                '提前性能测试，优化设计',
                '建立测试环境，提前验证',
                '建立SLA，定期检查，备选供应商',
                '做迁移测试，备份方案，分阶段迁移',
                '原型演示，定期评审，对齐期望'
            ],
            '责任人': ['技术总监', '项目经理', '产品经理', '技术负责人',
                      '测试负责人', '技术负责人', '技术负责人', '项目经理',
                      '技术负责人', '项目经理'],
            '状态': ['监控中'] * 10,
            '触发事件': [''] * 10,  # 空字符串而非NaN
            '更新日期': [pd.Timestamp(datetime.now())] * 10
        }

        df = pd.DataFrame(data)
        df.to_excel(output_file, index=False)
        self._format_risk_register(output_file)
        print(f"✓ 风险登记册已创建: {output_file}")
        return df

    def _format_risk_register(self, filepath):
        """格式化风险登记册"""
        book = load_workbook(filepath)
        sheet = book.active

        # 设置标题样式
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 根据风险等级设置颜色
        for row in sheet.iter_rows(min_row=2):
            risk_level = row[6].value  # 风险等级列
            if risk_level == '高':
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif risk_level == '中':
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            for cell in row:
                if cell.column <= 6:  # 只给前6列着色
                    cell.fill = fill

        # 自动列宽
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        chinese_chars = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                        cell_length = cell_length + chinese_chars
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max(max_length + 2, 10) * 1.1
            sheet.column_dimensions[column].width = min(adjusted_width, 40)

        # 冻结首行
        sheet.freeze_panes = 'A2'

        book.save(filepath)
        print(f"✓ 风险登记册已格式化")

    def add_risk(self, filepath, risk_data):
        """
        添加新风险

        参数:
            filepath: 风险登记册文件路径
            risk_data: 风险数据字典
                {
                    '风险描述': '...',
                    '风险类别': '技术/资源/时间/范围/质量/外部',
                    '可能性(1-3)': 1-3,
                    '影响程度(1-3)': 1-3,
                    '应对策略': '规避/缓解/转移/接受',
                    '应对措施': '...',
                    '责任人': '...'
                }
        """
        df = pd.read_excel(filepath)
        # 确保触发事件列是字符串类型
        df['触发事件'] = df['触发事件'].fillna('').astype(str)

        # 生成新的风险ID
        max_id = df['风险ID'].str.extract(r'R(\d+)').astype(int).max().values[0]
        new_id = f'R{max_id + 1:03d}'

        # 计算风险等级
        likelihood = int(risk_data['可能性(1-3)'])
        impact = int(risk_data['影响程度(1-3)'])
        score = likelihood * impact

        if score >= 7:
            risk_level = '高'
        elif score >= 4:
            risk_level = '中'
        else:
            risk_level = '低'

        # 创建新风险
        new_risk = pd.DataFrame({
            '风险ID': [new_id],
            '风险描述': [risk_data['风险描述']],
            '风险类别': [risk_data['风险类别']],
            '可能性(1-3)': [likelihood],
            '影响程度(1-3)': [impact],
            '风险等级': [risk_level],
            '应对策略': [risk_data['应对策略']],
            '应对措施': [risk_data['应对措施']],
            '责任人': [risk_data['责任人']],
            '状态': ['监控中'],
            '触发事件': [''],
            '更新日期': [pd.Timestamp(datetime.now())],
            '触发事件': ['']
        })

        # 追加到现有数据
        df = pd.concat([df, new_risk], ignore_index=True)
        df.to_excel(filepath, index=False)
        self._format_risk_register(filepath)

        print(f"✓ 新风险已添加: {new_id}")
        print(f"  - 风险等级: {risk_level} ({likelihood}×{impact}={score})")
        return new_id

    def update_risk_status(self, filepath, risk_id, new_status, trigger_event=''):
        """
        更新风险状态

        参数:
            filepath: 风险登记册文件路径
            risk_id: 风险ID (如 'R001')
            new_status: 新状态 (监控中/已缓解/已关闭/已触发)
            trigger_event: 触发事件描述 (如已触发)
        """
        df = pd.read_excel(filepath)
        # 确保触发事件列是字符串类型
        df['触发事件'] = df['触发事件'].fillna('').astype(str)

        # 查找风险
        idx = df[df['风险ID'] == risk_id].index
        if len(idx) == 0:
            print(f"✗ 未找到风险: {risk_id}")
            return

        # 更新状态
        df.loc[idx, '状态'] = new_status
        df.loc[idx, '更新日期'] = pd.Timestamp(datetime.now())
        if trigger_event:
            df.loc[idx, '触发事件'] = trigger_event

        df.to_excel(filepath, index=False)
        self._format_risk_register(filepath)

        print(f"✓ 风险状态已更新: {risk_id} → {new_status}")
        if trigger_event:
            print(f"  - 触发事件: {trigger_event}")

    def analyze_risks(self, filepath):
        """
        分析风险

        返回:
            分析报告 (字典)
        """
        df = pd.read_excel(filepath)

        analysis = {
            '总风险数': len(df),
            '按类别分布': df.groupby('风险类别').size().to_dict(),
            '按等级分布': df.groupby('风险等级').size().to_dict(),
            '按状态分布': df.groupby('状态').size().to_dict(),
            '高风险风险': df[df['风险等级'] == '高'].to_dict('records'),
            '已触发风险': df[df['状态'] == '已触发'].to_dict('records'),
            '监控中风险': df[df['状态'] == '监控中'].to_dict('records'),
            '平均影响': df['影响程度(1-3)'].mean(),
            '平均可能性': df['可能性(1-3)'].mean()
        }

        return analysis

    def generate_risk_report(self, filepath, output_file='风险报告.md'):
        """
        生成风险报告

        参数:
            filepath: 风险登记册文件路径
            output_file: 输出报告文件路径
        """
        df = pd.read_excel(filepath)
        analysis = self.analyze_risks(filepath)

        report = []
        report.append("# 📊 风险管理报告")
        report.append(f"\n生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # 概览
        report.append("\n## 📈 风险概览")
        report.append(f"- **总风险数**: {analysis['总风险数']}")
        report.append(f"- **高风险**: {analysis['按等级分布'].get('高', 0)}")
        report.append(f"- **中风险**: {analysis['按等级分布'].get('中', 0)}")
        report.append(f"- **低风险**: {analysis['按等级分布'].get('低', 0)}")
        report.append(f"- **监控中**: {analysis['按状态分布'].get('监控中', 0)}")
        report.append(f"- **已触发**: {analysis['按状态分布'].get('已触发', 0)}")
        report.append(f"- **平均可能性**: {analysis['平均可能性']:.1f}")
        report.append(f"- **平均影响**: {analysis['平均影响']:.1f}")

        # 按类别分布
        report.append("\n## 📋 按类别分布")
        for category, count in analysis['按类别分布'].items():
            report.append(f"- **{category}**: {count}")

        # 高风险列表
        high_risks = df[df['风险等级'] == '高']
        if not high_risks.empty:
            report.append("\n## 🔴 高风险风险")
            for _, row in high_risks.iterrows():
                report.append(f"\n### {row['风险ID']}: {row['风险描述']}")
                report.append(f"- **类别**: {row['风险类别']}")
                report.append(f"- **可能性**: {row['可能性(1-3)']}/3")
                report.append(f"- **影响**: {row['影响程度(1-3)']}/3")
                report.append(f"- **应对策略**: {row['应对策略']}")
                report.append(f"- **责任人**: {row['责任人']}")
                report.append(f"- **状态**: {row['状态']}")
                report.append(f"- **应对措施**: {row['应对措施']}")

        # 已触发风险
        triggered = df[df['状态'] == '已触发']
        if not triggered.empty:
            report.append("\n## ⚠️ 已触发风险")
            for _, row in triggered.iterrows():
                report.append(f"\n### {row['风险ID']}: {row['风险描述']}")
                report.append(f"- **触发事件**: {row['触发事件']}")
                report.append(f"- **应对策略**: {row['应对策略']}")
                report.append(f"- **应对措施**: {row['应对措施']}")

        # 监控中风险
        monitoring = df[df['状态'] == '监控中']
        if not monitoring.empty:
            report.append("\n## 👀 监控中风险")
            for _, row in monitoring.iterrows():
                status_icon = '🟢' if row['风险等级'] == '低' else '🟡' if row['风险等级'] == '中' else '🔴'
                report.append(f"\n{status_icon} **{row['风险ID']}**: {row['风险描述']}")
                report.append(f"   - 等级: {row['风险等级']} | 类别: {row['风险类别']}")
                report.append(f"   - 责任人: {row['责任人']} | 策略: {row['应对策略']}")

        # 风险趋势建议
        report.append("\n## 💡 建议")

        # 检查高风险数量
        high_risk_count = analysis['按等级分布'].get('高', 0)
        if high_risk_count > 3:
            report.append("- ⚠️ 高风险数量较多，建议优先制定应对措施")

        # 检查已触发风险
        if analysis['按状态分布'].get('已触发', 0) > 0:
            report.append("- ⚠️ 已有风险触发，建议立即采取应对措施")

        # 检查监控中风险
        monitoring_count = analysis['按状态分布'].get('监控中', 0)
        if monitoring_count > 10:
            report.append("- 📊 监控中风险较多，建议定期检查")

        # 检查风险类别
        if '资源' in analysis['按类别分布'] and analysis['按类别分布']['资源'] > 3:
            report.append("- 👥 资源类风险较多，建议加强人员管理")

        if '技术' in analysis['按类别分布'] and analysis['按类别分布']['技术'] > 3:
            report.append("- 🔧 技术类风险较多，建议加强技术调研")

        if '范围' in analysis['按类别分布'] and analysis['按类别分布']['范围'] > 3:
            report.append("- 📏 范围类风险较多，建议建立变更流程")

        # 保存报告
        report_text = '\n'.join(report)
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(report_text)

        print(f"✓ 风险报告已生成: {output_file}")
        return report_text

    def export_to_json(self, filepath, output_file='风险数据.json'):
        """导出风险数据为JSON格式"""
        df = pd.read_excel(filepath)
        df.to_json(output_file, orient='records', force_ascii=False, indent=2)
        print(f"✓ 风险数据已导出: {output_file}")
        return output_file

    def import_from_json(self, filepath, json_file):
        """从JSON导入风险数据"""
        df_existing = pd.read_excel(filepath)
        df_new = pd.read_json(json_file)

        # 合并数据
        df = pd.concat([df_existing, df_new], ignore_index=True)
        df.to_excel(filepath, index=False)
        self._format_risk_register(filepath)

        print(f"✓ 风险数据已导入: {json_file}")


class RiskAssessmentHelper:
    """风险评估辅助工具"""

    @staticmethod
    def assess_likelihood(description):
        """
        评估可能性

        参数:
            description: 风险描述

        返回:
            可能性等级 (1-3)
        """
        # 高可能性关键词
        high_keywords = ['经常', '总是', '普遍', '频繁', '每次', '必然']

        # 中可能性关键词
        medium_keywords = ['偶尔', '有时', '可能', '也许', '或']

        # 低可能性关键词
        low_keywords = ['罕见', '少见', '几乎不', '不太可能', '很少']

        desc_lower = description.lower()

        if any(kw in desc_lower for kw in high_keywords):
            return 3
        elif any(kw in desc_lower for kw in medium_keywords):
            return 2
        elif any(kw in desc_lower for kw in low_keywords):
            return 1
        else:
            return 2  # 默认中等

    @staticmethod
    def assess_impact(description):
        """
        评估影响程度

        参数:
            description: 风险描述

        返回:
            影响程度等级 (1-3)
        """
        # 高影响关键词
        high_keywords = ['失败', '崩溃', '严重', '致命', '灾难',
                        '无法恢复', '延期超过30%', '项目失败', '数据丢失']

        # 中影响关键词
        medium_keywords = ['延期', '返工', '成本增加', '性能下降',
                          '部分功能', '影响范围中等', '成本增加10-30%']

        # 低影响关键词
        low_keywords = ['轻微', '小', '影响小', '容易修复',
                       '成本低', '影响范围小', '快速恢复']

        desc_lower = description.lower()

        if any(kw in desc_lower for kw in high_keywords):
            return 3
        elif any(kw in desc_lower for kw in medium_keywords):
            return 2
        elif any(kw in desc_lower for kw in low_keywords):
            return 1
        else:
            return 2  # 默认中等

    @staticmethod
    def suggest_strategy(likelihood, impact):
        """
        建议应对策略

        参数:
            likelihood: 可能性 (1-3)
            impact: 影响 (1-3)

        返回:
            建议策略 (规避/缓解/转移/接受)
        """
        score = likelihood * impact

        if score >= 7:
            return '规避'  # 高风险
        elif score >= 4:
            return '缓解'  # 中风险
        else:
            return '接受'  # 低风险


def main():
    """测试代码"""
    print("=== 项目风险管理工具 ===\n")

    # 创建风险管理器
    rm = RiskManager()

    # 创建风险登记册
    print("1. 创建风险登记册...")
    df = rm.create_risk_register('risk_register.xlsx')

    # 添加新风险
    print("\n2. 添加新风险...")
    new_risk = {
        '风险描述': '服务器可能宕机，导致系统不可用',
        '风险类别': '技术',
        '可能性(1-3)': 2,
        '影响程度(1-3)': 3,
        '应对策略': '缓解',
        '应对措施': '做负载均衡，监控告警，灾备方案',
        '责任人': '技术总监'
    }
    rm.add_risk('risk_register.xlsx', new_risk)

    # 生成报告
    print("\n3. 生成风险报告...")
    report = rm.generate_risk_report('risk_register.xlsx', 'risk_report.md')
    print("\n" + report[:500] + "...")

    # 分析风险
    print("\n4. 分析风险...")
    analysis = rm.analyze_risks('risk_register.xlsx')
    print(f"总风险数: {analysis['总风险数']}")
    print(f"按等级分布: {analysis['按等级分布']}")
    print(f"按状态分布: {analysis['按状态分布']}")

    print("\n✓ 测试完成!")


if __name__ == '__main__':
    main()
